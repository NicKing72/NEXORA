"""Safe tabular inspection and parsing for CSV and Excel sources."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from xlrd import open_workbook

from nexora_api.core.exceptions import DataStudioError


@dataclass
class TabularData:
    frame: pd.DataFrame
    available_sheets: list[str]
    selected_sheet: str | None
    duplicate_columns: list[str]


def inspect_excel_sheets(path: Path, file_type: str) -> list[str]:
    try:
        if file_type == "xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                return list(workbook.sheetnames)
            finally:
                workbook.close()
        if file_type == "xls":
            workbook = open_workbook(path, on_demand=True)
            try:
                return workbook.sheet_names()
            finally:
                workbook.release_resources()
    except Exception as exc:
        raise DataStudioError("corrupt_excel", "The Excel workbook could not be read.") from exc
    return []


def _duplicate_names(names: list[object]) -> list[str]:
    normalized = [str(name).strip() for name in names if name is not None]
    return sorted({name for name in normalized if normalized.count(name) > 1})


def _csv_headers(path: Path) -> tuple[list[str], str]:
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as source:
                sample = source.read(8192)
                if not sample.strip():
                    return [], encoding
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                source.seek(0)
                return next(csv.reader(source, dialect), []), encoding
        except (UnicodeDecodeError, csv.Error):
            continue
    raise DataStudioError("corrupt_csv", "The CSV file could not be decoded or parsed.")


def _excel_headers(path: Path, file_type: str, sheet: str) -> list[object]:
    if file_type == "xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[sheet]
            return list(next(worksheet.iter_rows(values_only=True), ()))
        finally:
            workbook.close()
    workbook = open_workbook(path, on_demand=True)
    try:
        worksheet = workbook.sheet_by_name(sheet)
        return worksheet.row_values(0) if worksheet.nrows else []
    finally:
        workbook.release_resources()


def parse_tabular(path: Path, file_type: str, selected_sheet: str | None = None) -> TabularData:
    """Parse source values without executing formulas or altering business data."""
    try:
        if file_type == "csv":
            headers, encoding = _csv_headers(path)
            frame = pd.read_csv(path, encoding=encoding, sep=None, engine="python")
            sheets: list[str] = []
            chosen_sheet = None
            duplicates = _duplicate_names(headers)
        else:
            sheets = inspect_excel_sheets(path, file_type)
            if not sheets:
                raise DataStudioError("no_sheets", "The workbook does not contain any sheets.")
            chosen_sheet = selected_sheet or sheets[0]
            if chosen_sheet not in sheets:
                raise DataStudioError("invalid_sheet", "The selected Excel sheet does not exist.")
            headers = _excel_headers(path, file_type, chosen_sheet)
            engine = "openpyxl" if file_type == "xlsx" else "xlrd"
            frame = pd.read_excel(path, sheet_name=chosen_sheet, engine=engine)
            duplicates = _duplicate_names(headers)
    except DataStudioError:
        raise
    except (ValueError, OSError, pd.errors.ParserError) as exc:
        raise DataStudioError("corrupt_file", "The tabular file could not be parsed.") from exc

    if len(frame.columns) == 0:
        raise DataStudioError("no_columns", "The dataset does not contain columns.")
    if frame.empty:
        raise DataStudioError("no_rows", "The dataset does not contain data rows.")

    frame.columns = [
        str(column).strip() or f"unnamed_{index + 1}" for index, column in enumerate(frame.columns)
    ]
    return TabularData(
        frame=frame,
        available_sheets=sheets,
        selected_sheet=chosen_sheet,
        duplicate_columns=duplicates,
    )


def json_safe_value(value: object) -> object:
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.isoformat()
    if hasattr(value, "item"):
        return value.item()  # type: ignore[no-any-return]
    return value


def preview_records(frame: pd.DataFrame, limit: int) -> list[dict[str, object]]:
    return [
        {str(column): json_safe_value(value) for column, value in row.items()}
        for row in frame.head(limit).to_dict(orient="records")
    ]
